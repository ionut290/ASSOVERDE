from __future__ import annotations

import sqlite3
from contextlib import closing
from pathlib import Path
from tempfile import NamedTemporaryFile
from typing import List, Optional

from fastapi import FastAPI, File, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from openpyxl import load_workbook
from pydantic import BaseModel, Field

BASE_DIR = Path(__file__).resolve().parent.parent
DB_PATH = BASE_DIR / "assoverde.db"

app = FastAPI(title="Assoverde Preventivi API", version="0.2.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


class PricelistItemIn(BaseModel):
    codice_prezzo: str
    capitolo: str
    descrizione: str
    unita_misura: str
    prezzo_unitario: float = Field(ge=0)


class PricelistItem(PricelistItemIn):
    id: int


class UploadResult(BaseModel):
    inserted: int
    skipped: int


class SuggestRequest(BaseModel):
    descrizione_lavoro: str
    top_k: int = Field(default=5, ge=1, le=20)


class QuoteLineIn(BaseModel):
    item_id: int
    quantita: float = Field(gt=0)


class QuoteCreateRequest(BaseModel):
    cliente: str
    oggetto: str
    descrizione_lavoro: str
    righe: List[QuoteLineIn]


class QuoteLineOut(BaseModel):
    descrizione: str
    unita_misura: str
    prezzo_unitario: float
    quantita: float
    subtotale: float


class QuoteOut(BaseModel):
    id: int
    cliente: str
    oggetto: str
    descrizione_lavoro: str
    righe: List[QuoteLineOut]
    totale: float


HEADER_ALIASES = {
    "codice_prezzo": {"codice_prezzo", "codice", "codice prezzo"},
    "capitolo": {"capitolo"},
    "descrizione": {"descrizione", "voce"},
    "unita_misura": {"unita_misura", "unita di misura", "unità di misura", "um"},
    "prezzo_unitario": {"prezzo_unitario", "prezzo orario", "prezzo", "prezzo unitario"},
}


def normalize(value: str) -> str:
    return " ".join(value.strip().lower().replace("_", " ").split())


def extract_excel_rows(file_path: Path) -> list[dict[str, str | float]]:
    workbook = load_workbook(filename=file_path, data_only=True)
    sheet = workbook.active

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [normalize(str(cell or "")) for cell in rows[0]]
    positions: dict[str, int] = {}

    for idx, header in enumerate(headers):
        for canonical, aliases in HEADER_ALIASES.items():
            if header in aliases:
                positions[canonical] = idx

    missing = [key for key in HEADER_ALIASES if key not in positions]
    if missing:
        raise HTTPException(
            status_code=400,
            detail=(
                "Colonne mancanti nel file Excel. Richieste: "
                "codice_prezzo, capitolo, descrizione, unita_misura, prezzo_unitario"
            ),
        )

    parsed: list[dict[str, str | float]] = []
    for row in rows[1:]:
        codice = str(row[positions["codice_prezzo"]] or "").strip()
        descrizione = str(row[positions["descrizione"]] or "").strip()
        if not codice or not descrizione:
            continue

        raw_price = row[positions["prezzo_unitario"]]
        try:
            prezzo = float(raw_price)
        except (TypeError, ValueError):
            continue

        parsed.append(
            {
                "codice_prezzo": codice,
                "capitolo": str(row[positions["capitolo"]] or "").strip() or "N/D",
                "descrizione": descrizione,
                "unita_misura": str(row[positions["unita_misura"]] or "").strip() or "ora",
                "prezzo_unitario": prezzo,
            }
        )

    return parsed


def get_conn() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db() -> None:
    with closing(get_conn()) as conn:
        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS pricelist_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                codice_prezzo TEXT NOT NULL UNIQUE,
                capitolo TEXT NOT NULL,
                descrizione TEXT NOT NULL,
                unita_misura TEXT NOT NULL,
                prezzo_unitario REAL NOT NULL
            );

            CREATE TABLE IF NOT EXISTS quotes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                cliente TEXT NOT NULL,
                oggetto TEXT NOT NULL,
                descrizione_lavoro TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS quote_lines (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                quote_id INTEGER NOT NULL,
                item_id INTEGER NOT NULL,
                quantita REAL NOT NULL,
                prezzo_unitario REAL NOT NULL,
                FOREIGN KEY(quote_id) REFERENCES quotes(id),
                FOREIGN KEY(item_id) REFERENCES pricelist_items(id)
            );
            """
        )
        conn.commit()


@app.on_event("startup")
def on_startup() -> None:
    init_db()


@app.get("/health")
def health() -> dict[str, str]:
    return {"status": "ok"}


@app.post("/pricelist/items", response_model=PricelistItem)
def create_pricelist_item(payload: PricelistItemIn) -> PricelistItem:
    with closing(get_conn()) as conn:
        try:
            cursor = conn.execute(
                """
                INSERT INTO pricelist_items
                (codice_prezzo, capitolo, descrizione, unita_misura, prezzo_unitario)
                VALUES (?, ?, ?, ?, ?)
                """,
                (
                    payload.codice_prezzo,
                    payload.capitolo,
                    payload.descrizione,
                    payload.unita_misura,
                    payload.prezzo_unitario,
                ),
            )
            conn.commit()
        except sqlite3.IntegrityError as exc:
            raise HTTPException(status_code=409, detail="Codice prezzo già presente") from exc

    return PricelistItem(id=cursor.lastrowid, **payload.model_dump())


@app.post("/pricelist/upload", response_model=UploadResult)
async def upload_pricelist_excel(file: UploadFile = File(...)) -> UploadResult:
    if not file.filename:
        raise HTTPException(status_code=400, detail="File non valido")

    extension = Path(file.filename).suffix.lower()
    if extension not in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        raise HTTPException(status_code=400, detail="Carica un file Excel (.xlsx)")

    with NamedTemporaryFile(delete=False, suffix=extension) as tmp:
        content = await file.read()
        tmp.write(content)
        temp_path = Path(tmp.name)

    parsed_rows = extract_excel_rows(temp_path)
    temp_path.unlink(missing_ok=True)

    inserted = 0
    skipped = 0
    with closing(get_conn()) as conn:
        for row in parsed_rows:
            try:
                conn.execute(
                    """
                    INSERT INTO pricelist_items
                    (codice_prezzo, capitolo, descrizione, unita_misura, prezzo_unitario)
                    VALUES (?, ?, ?, ?, ?)
                    """,
                    (
                        row["codice_prezzo"],
                        row["capitolo"],
                        row["descrizione"],
                        row["unita_misura"],
                        row["prezzo_unitario"],
                    ),
                )
                inserted += 1
            except sqlite3.IntegrityError:
                skipped += 1

        conn.commit()

    return UploadResult(inserted=inserted, skipped=skipped)


@app.get("/pricelist/items", response_model=List[PricelistItem])
def list_pricelist_items(search: Optional[str] = None) -> List[PricelistItem]:
    query = "SELECT * FROM pricelist_items"
    params: list[str] = []

    if search:
        query += " WHERE lower(codice_prezzo) LIKE ? OR lower(descrizione) LIKE ? OR lower(capitolo) LIKE ?"
        pattern = f"%{search.lower()}%"
        params.extend([pattern, pattern, pattern])

    query += " ORDER BY id DESC"

    with closing(get_conn()) as conn:
        rows = conn.execute(query, params).fetchall()

    return [PricelistItem(**dict(row)) for row in rows]


def score_item(text: str, item: sqlite3.Row) -> int:
    score = 0
    haystack = f"{item['descrizione']} {item['capitolo']} {item['codice_prezzo']}".lower()
    for token in text.lower().split():
        if token in haystack:
            score += 1
    return score


@app.post("/quotes/suggest", response_model=List[PricelistItem])
def suggest_items(payload: SuggestRequest) -> List[PricelistItem]:
    with closing(get_conn()) as conn:
        rows = conn.execute("SELECT * FROM pricelist_items").fetchall()

    ranked = sorted(rows, key=lambda r: score_item(payload.descrizione_lavoro, r), reverse=True)
    non_zero = [row for row in ranked if score_item(payload.descrizione_lavoro, row) > 0]
    result = non_zero[: payload.top_k] if non_zero else ranked[: payload.top_k]
    return [PricelistItem(**dict(row)) for row in result]


@app.post("/quotes", response_model=QuoteOut)
def create_quote(payload: QuoteCreateRequest) -> QuoteOut:
    with closing(get_conn()) as conn:
        quote_cursor = conn.execute(
            "INSERT INTO quotes (cliente, oggetto, descrizione_lavoro) VALUES (?, ?, ?)",
            (payload.cliente, payload.oggetto, payload.descrizione_lavoro),
        )
        quote_id = quote_cursor.lastrowid

        out_lines: list[QuoteLineOut] = []
        total = 0.0

        for line in payload.righe:
            item = conn.execute("SELECT * FROM pricelist_items WHERE id = ?", (line.item_id,)).fetchone()
            if item is None:
                raise HTTPException(status_code=404, detail=f"Voce {line.item_id} non trovata")

            subtotal = float(item["prezzo_unitario"]) * line.quantita
            total += subtotal

            conn.execute(
                """
                INSERT INTO quote_lines (quote_id, item_id, quantita, prezzo_unitario)
                VALUES (?, ?, ?, ?)
                """,
                (quote_id, line.item_id, line.quantita, float(item["prezzo_unitario"])),
            )

            out_lines.append(
                QuoteLineOut(
                    descrizione=item["descrizione"],
                    unita_misura=item["unita_misura"],
                    prezzo_unitario=float(item["prezzo_unitario"]),
                    quantita=line.quantita,
                    subtotale=subtotal,
                )
            )

        conn.commit()

    return QuoteOut(
        id=quote_id,
        cliente=payload.cliente,
        oggetto=payload.oggetto,
        descrizione_lavoro=payload.descrizione_lavoro,
        righe=out_lines,
        totale=round(total, 2),
    )
